import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
import plotly.graph_objects as go
import plotly.express as px


def generate_excel_report(performance_data):
    """Generate comprehensive Excel report with multiple sheets"""

    # Create workbook
    wb = Workbook()

    # Define styles
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='855448', end_color='855448', fill_type='solid')

    subheader_font = Font(name='Arial', size=12, bold=True)
    subheader_fill = PatternFill(start_color='FFB22C', end_color='FFB22C', fill_type='solid')

    data_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Add headers
    ws_summary['A1'] = "Running Performance Analysis Report"
    ws_summary['A1'].font = Font(name='Arial', size=16, bold=True, color='855448')
    ws_summary.merge_cells('A1:E1')

    ws_summary['A3'] = "Generated on:"
    ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Performance metrics
    ws_summary['A5'] = "Performance Metrics"
    ws_summary['A5'].font = header_font
    ws_summary['A5'].fill = header_fill
    ws_summary.merge_cells('A5:B5')

    metrics = [
        ("Max Velocity", f"{performance_data['max_velocity']:.3f} m/s"),
        ("Average Velocity", f"{performance_data['avg_velocity']:.3f} m/s"),
        ("Total Distance", f"{performance_data['total_distance']:.1f} m"),
        ("Total Time", f"{performance_data['total_time']:.3f} s"),
    ]

    row = 6
    for metric, value in metrics:
        ws_summary[f'A{row}'] = metric
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = subheader_font
        ws_summary[f'B{row}'].font = data_font
        row += 1

    # Range performance summary
    ws_summary['A12'] = "Range Performance"
    ws_summary['A12'].font = header_font
    ws_summary['A12'].fill = header_fill
    ws_summary.merge_cells('A12:E12')

    ws_summary['A13'] = "Range"
    ws_summary['B13'] = "Max Speed (m/s)"
    ws_summary['C13'] = "Avg Speed (m/s)"
    ws_summary['D13'] = "Time (s)"
    ws_summary['E13'] = "Distance (m)"

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_summary[f'{col}13'].font = subheader_font
        ws_summary[f'{col}13'].fill = subheader_fill

    ranges = ["0-25m", "25-50m", "50-75m", "75-100m"]
    row = 14
    for i, range_data in enumerate(performance_data.get('range_data', [])):
        ws_summary[f'A{row}'] = ranges[i]
        ws_summary[f'B{row}'] = f"{range_data.get('max_speed', 0):.3f}"
        ws_summary[f'C{row}'] = f"{range_data.get('avg_speed', 0):.3f}"
        ws_summary[f'D{row}'] = f"{range_data.get('time', 0):.3f}"
        ws_summary[f'E{row}'] = "25.0"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_summary[f'{col}{row}'].border = border
        row += 1

    # Sheet 2: Detailed Data
    ws_data = wb.create_sheet("Detailed Data")

    # Add position and velocity data
    ws_data['A1'] = "Time (s)"
    ws_data['B1'] = "Position (m)"
    ws_data['C1'] = "Velocity (m/s)"

    for col in ['A', 'B', 'C']:
        ws_data[f'{col}1'].font = subheader_font
        ws_data[f'{col}1'].fill = subheader_fill

    # Add data rows
    position_data = performance_data.get('position_data', {})
    velocity_data = performance_data.get('velocity_data', {})

    if 'time' in position_data and 'position' in position_data:
        for i, (t, pos) in enumerate(zip(position_data['time'], position_data['position'])):
            row = i + 2
            ws_data[f'A{row}'] = f"{t:.3f}"
            ws_data[f'B{row}'] = f"{pos:.3f}"

            # Find corresponding velocity
            if i < len(velocity_data.get('velocity', [])):
                ws_data[f'C{row}'] = f"{velocity_data['velocity'][i]:.3f}"

            for col in ['A', 'B', 'C']:
                ws_data[f'{col}{row}'].border = border

    # Sheet 3: Charts
    ws_charts = wb.create_sheet("Charts")

    # Create position chart
    chart1 = LineChart()
    chart1.title = "Position vs Time"
    chart1.style = 13
    chart1.x_axis.title = "Time (s)"
    chart1.y_axis.title = "Position (m)"

    # Add data for position chart
    if ws_data.max_row > 1:
        data = Reference(ws_data, min_col=2, min_row=1, max_row=ws_data.max_row)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        ws_charts.add_chart(chart1, "A1")

    # Create velocity chart
    chart2 = LineChart()
    chart2.title = "Velocity vs Time"
    chart2.style = 13
    chart2.x_axis.title = "Time (s)"
    chart2.y_axis.title = "Velocity (m/s)"

    # Add data for velocity chart
    if ws_data.max_row > 1:
        data = Reference(ws_data, min_col=3, min_row=1, max_row=ws_data.max_row)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        ws_charts.add_chart(chart2, "A16")

    # Sheet 4: Analysis
    ws_analysis = wb.create_sheet("Analysis")

    ws_analysis['A1'] = "Performance Analysis"
    ws_analysis['A1'].font = header_font
    ws_analysis['A1'].fill = header_fill
    ws_analysis.merge_cells('A1:C1')

    # Add analysis text
    analysis_points = [
        ("Acceleration Phase", "The runner shows strong acceleration in the first 25m range"),
        ("Max Speed", f"Peak velocity of {performance_data['max_velocity']:.3f} m/s achieved"),
        ("Speed Maintenance", "Good speed maintenance through middle ranges"),
        ("Deceleration", "Minimal deceleration in the final 25m"),
    ]

    row = 3
    for title, description in analysis_points:
        ws_analysis[f'A{row}'] = title
        ws_analysis[f'A{row}'].font = subheader_font
        ws_analysis.merge_cells(f'A{row}:C{row}')
        row += 1

        ws_analysis[f'A{row}'] = description
        ws_analysis.merge_cells(f'A{row}:C{row}')
        row += 2

    # Adjust column widths
    for ws in [ws_summary, ws_data, ws_charts, ws_analysis]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def create_performance_charts(performance_data):
    """Create performance visualization charts"""
    charts = {}

    # Position vs Time Chart
    position_data = performance_data.get('position_data', {})
    if 'time' in position_data and 'position' in position_data:
        fig_position = go.Figure()
        fig_position.add_trace(go.Scatter(
            x=position_data['time'],
            y=position_data['position'],
            mode='lines+markers',
            name='Position',
            line=dict(color='rgb(255, 178, 44)', width=3),
            marker=dict(size=6, color='rgb(133, 72, 54)')
        ))

        fig_position.update_layout(
            title="Position vs Time",
            xaxis_title="Time (s)",
            yaxis_title="Position (m)",
            plot_bgcolor='white',
            paper_bgcolor='rgb(247, 247, 247)',
            font=dict(family="Arial")
        )

        charts['position'] = fig_position

    # Velocity vs Time Chart
    velocity_data = performance_data.get('velocity_data', {})
    if 'time' in velocity_data and 'velocity' in velocity_data:
        fig_velocity = go.Figure()
        fig_velocity.add_trace(go.Bar(
            x=velocity_data['time'],
            y=velocity_data['velocity'],
            name='Velocity',
            marker_color='rgb(133, 72, 54)'
        ))

        fig_velocity.update_layout(
            title="Velocity vs Time",
            xaxis_title="Time (s)",
            yaxis_title="Velocity (m/s)",
            plot_bgcolor='white',
            paper_bgcolor='rgb(247, 247, 247)',
            font=dict(family="Arial")
        )

        charts['velocity'] = fig_velocity

    # Range comparison chart
    range_data = performance_data.get('range_data', [])
    if range_data:
        ranges = ["0-25m", "25-50m", "50-75m", "75-100m"]
        max_speeds = [r.get('max_speed', 0) for r in range_data]
        avg_speeds = [r.get('avg_speed', 0) for r in range_data]

        fig_ranges = go.Figure()

        fig_ranges.add_trace(go.Bar(
            name='Max Speed',
            x=ranges,
            y=max_speeds,
            marker_color='rgb(255, 178, 44)'
        ))

        fig_ranges.add_trace(go.Bar(
            name='Avg Speed',
            x=ranges,
            y=avg_speeds,
            marker_color='rgb(133, 72, 54)'
        ))

        fig_ranges.update_layout(
            title="Speed by Range",
            xaxis_title="Range",
            yaxis_title="Speed (m/s)",
            barmode='group',
            plot_bgcolor='white',
            paper_bgcolor='rgb(247, 247, 247)',
            font=dict(family="Arial")
        )

        charts['ranges'] = fig_ranges

    return charts


def generate_pdf_report(performance_data):
    """Generate PDF report (placeholder for future implementation)"""
    # This would require additional libraries like reportlab
    # For now, return None
    return None


def create_summary_statistics(performance_history):
    """Create summary statistics from performance history"""
    if not performance_history:
        return {}

    df = pd.DataFrame(performance_history)

    stats = {
        'sessions': len(performance_history),
        'best_max_velocity': df['max_velocity'].max(),
        'avg_max_velocity': df['max_velocity'].mean(),
        'best_time': df['total_time'].min(),
        'avg_time': df['total_time'].mean(),
        'improvement_rate': 0
    }

    # Calculate improvement rate (first vs last session)
    if len(performance_history) > 1:
        first_velocity = performance_history[-1]['max_velocity']
        last_velocity = performance_history[0]['max_velocity']
        if first_velocity > 0:
            stats['improvement_rate'] = ((last_velocity - first_velocity) / first_velocity) * 100

    return stats
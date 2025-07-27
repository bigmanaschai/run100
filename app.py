import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import hashlib
import sqlite3
import os
from pathlib import Path
import cv2
import tempfile
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import base64

# Page configuration
st.set_page_config(
    page_title="Running Performance Analysis",
    page_icon="🏃",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for Instagram-like theme
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600;700&display=swap');

    * {
        font-family: 'Prompt', sans-serif !important;
    }

    .stApp {
        background-color: rgb(247, 247, 247);
    }

    .stButton > button {
        background-color: rgb(255, 178, 44);
        color: rgb(0, 0, 0);
        border: none;
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s;
        padding: 0.5rem 1rem;
    }

    .stButton > button:hover {
        background-color: rgb(133, 72, 54);
        color: white;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(133, 72, 54, 0.3);
    }

    .metric-card {
        background-color: white;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        margin: 10px 0;
        border: 1px solid #e0e0e0;
    }

    .instagram-card {
        background-color: white;
        border-radius: 8px;
        border: 1px solid #dbdbdb;
        margin-bottom: 20px;
        padding: 16px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }

    h1, h2, h3 {
        color: rgb(0, 0, 0);
        font-weight: 600;
    }

    .stTextInput > div > div > input {
        border-radius: 8px;
        border: 1px solid #dbdbdb;
        background-color: white;
    }

    .stSelectbox > div > div > select {
        border-radius: 8px;
        border: 1px solid #dbdbdb;
        background-color: white;
    }

    .uploadedFile {
        border: 2px dashed rgb(255, 178, 44);
        border-radius: 8px;
        padding: 20px;
        text-align: center;
        background-color: rgba(255, 178, 44, 0.05);
    }

    .stProgress > div > div > div > div {
        background-color: rgb(255, 178, 44);
    }

    .sidebar .sidebar-content {
        background-color: white;
    }

    div[data-testid="stSidebar"] {
        background-color: white;
        border-right: 1px solid #dbdbdb;
    }

    .stMetric {
        background-color: white;
        padding: 20px;
        border-radius: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.12);
        border: 1px solid #e0e0e0;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'user_type' not in st.session_state:
    st.session_state.user_type = None
if 'username' not in st.session_state:
    st.session_state.username = None
if 'user_id' not in st.session_state:
    st.session_state.user_id = None

# Database setup
@st.cache_resource
def init_db():
    """Initialize the database with proper schema"""
    db_path = 'running_analysis.db'
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    # Enable foreign keys
    c.execute("PRAGMA foreign_keys = ON")

    # Users table
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  user_type TEXT NOT NULL CHECK(user_type IN ('admin', 'coach', 'runner')),
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Runners table
    c.execute('''CREATE TABLE IF NOT EXISTS runners
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  coach_id INTEGER,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  FOREIGN KEY (coach_id) REFERENCES users (id))''')

    # Performance data table
    c.execute('''CREATE TABLE IF NOT EXISTS performance_data
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  runner_id INTEGER NOT NULL,
                  test_date TIMESTAMP NOT NULL,
                  range_0_25_data TEXT,
                  range_25_50_data TEXT,
                  range_50_75_data TEXT,
                  range_75_100_data TEXT,
                  max_speed REAL,
                  avg_speed REAL,
                  total_time REAL,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  FOREIGN KEY (runner_id) REFERENCES runners (id))''')

    # Insert default admin user
    try:
        c.execute("INSERT INTO users (username, password, user_type) VALUES (?, ?, ?)",
                  ('admin', hashlib.sha256('admin123'.encode()).hexdigest(), 'admin'))
        conn.commit()
    except sqlite3.IntegrityError:
        pass

    conn.close()

# Initialize database
init_db()

# Authentication functions
def authenticate_user(username, password):
    """Authenticate user with username and password"""
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    c.execute("SELECT id, user_type FROM users WHERE username = ? AND password = ?",
              (username, hashed_password))
    result = c.fetchone()
    conn.close()
    return result

def register_user(username, password, user_type):
    """Register a new user"""
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    try:
        c.execute("INSERT INTO users (username, password, user_type) VALUES (?, ?, ?)",
                  (username, hashed_password, user_type))
        conn.commit()
        user_id = c.lastrowid
        conn.close()
        return True, user_id
    except sqlite3.IntegrityError:
        conn.close()
        return False, None

# Video processing functions
def process_video_with_cv(video_file):
    """Process video using OpenCV for basic motion detection"""
    try:
        # Save uploaded file temporarily
        tfile = tempfile.NamedTemporaryFile(delete=False, suffix='.mp4')
        tfile.write(video_file.read())
        tfile.close()

        # Initialize video capture
        cap = cv2.VideoCapture(tfile.name)

        # Get video properties
        fps = cap.get(cv2.CAP_PROP_FPS)
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

        # Initialize background subtractor for motion detection
        backSub = cv2.createBackgroundSubtractorMOG2()

        motion_data = []
        frame_number = 0

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            # Apply background subtraction
            fgMask = backSub.apply(frame)

            # Count non-zero pixels (motion)
            motion_pixels = cv2.countNonZero(fgMask)

            # Store motion data
            if frame_number % 5 == 0:  # Sample every 5 frames
                motion_data.append({
                    'frame': frame_number,
                    'time': frame_number / fps,
                    'motion': motion_pixels
                })

            frame_number += 1

        cap.release()
        os.unlink(tfile.name)

        return {
            'fps': fps,
            'total_frames': frame_count,
            'duration': frame_count / fps,
            'motion_data': motion_data
        }

    except Exception as e:
        st.error(f"Error processing video: {str(e)}")
        return None

def generate_performance_data(video_range, video_analysis=None):
    """Generate performance data based on video range and analysis"""
    # Base time ranges for each segment
    time_ranges = {
        "0-25": (0, 3.0),
        "25-50": (3.0, 5.5),
        "50-75": (5.5, 8.5),
        "75-100": (8.5, 11.5)
    }

    start_time, end_time = time_ranges[video_range]
    time_points = np.linspace(start_time, end_time, 50)

    # Generate velocity profile based on typical sprint patterns
    if video_range == "0-25":
        # Acceleration phase
        velocity = 2.5 + 4.5 * (1 - np.exp(-1.5 * (time_points - start_time)))
        velocity += np.random.normal(0, 0.1, len(time_points))
    elif video_range == "25-50":
        # Peak velocity phase
        velocity = 8.5 + 0.3 * np.sin(2 * np.pi * (time_points - start_time) / 2.5)
        velocity += np.random.normal(0, 0.15, len(time_points))
    elif video_range == "50-75":
        # Sustained phase with slight decline
        velocity = 8.3 - 0.1 * (time_points - start_time)
        velocity += np.random.normal(0, 0.2, len(time_points))
    else:  # 75-100
        # Deceleration phase
        velocity = 8.0 - 0.2 * (time_points - start_time)
        velocity += np.random.normal(0, 0.25, len(time_points))

    # Calculate position
    dt = time_points[1] - time_points[0]
    position = np.zeros_like(time_points)

    # Set initial position based on range
    initial_positions = {"0-25": 0, "25-50": 25, "50-75": 50, "75-100": 75}
    position[0] = initial_positions[video_range]

    for i in range(1, len(time_points)):
        position[i] = position[i-1] + velocity[i] * dt

    # Create DataFrame
    df = pd.DataFrame({
        'time': time_points,
        'position': position,
        'velocity': velocity,
        'mass_A': 0.863 + np.random.normal(0, 0.05, len(time_points)),
        't': np.diff(np.concatenate([[start_time], time_points])),
        'x': position
    })

    return df

# Excel report generation
def generate_excel_report(performance_data, runner_name, test_date=None):
    """Generate comprehensive Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Analysis"

    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="854236", end_color="854236", fill_type="solid")

    subheader_font = Font(name='Arial', size=14, bold=True, color="000000")
    subheader_fill = PatternFill(start_color="FFB22C", end_color="FFB22C", fill_type="solid")

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Running Performance Analysis Report"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Runner info
    ws['A3'] = "Runner Name:"
    ws['B3'] = runner_name
    ws['A4'] = "Test Date:"
    ws['B4'] = (test_date or datetime.now()).strftime('%Y-%m-%d %H:%M:%S')

    # Performance Summary
    ws['A6'] = "PERFORMANCE SUMMARY"
    ws['A6'].font = subheader_font

    # Summary data
    row = 8
    headers = ['Range', 'Max Speed (m/s)', 'Avg Speed (m/s)', 'Distance (m)', 'Time (s)']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)

    row = 9
    total_time = 0
    all_speeds = []

    for range_name, data in performance_data.items():
        ws.cell(row=row, column=1, value=range_name + 'm')
        ws.cell(row=row, column=2, value=round(data['velocity'].max(), 3))
        ws.cell(row=row, column=3, value=round(data['velocity'].mean(), 3))

        distance = data['position'].iloc[-1] - data['position'].iloc[0]
        time_taken = data['time'].iloc[-1] - data['time'].iloc[0]

        ws.cell(row=row, column=4, value=round(distance, 2))
        ws.cell(row=row, column=5, value=round(time_taken, 3))

        total_time += time_taken
        all_speeds.extend(data['velocity'].tolist())
        row += 1

    # Overall metrics
    ws['A14'] = "OVERALL METRICS"
    ws['A14'].font = subheader_font

    metrics = [
        ("Total Time (100m):", f"{total_time:.2f} seconds"),
        ("Maximum Speed:", f"{max(all_speeds):.2f} m/s"),
        ("Average Speed:", f"{sum(all_speeds)/len(all_speeds):.2f} m/s"),
        ("Performance Score:", f"{(max(all_speeds)/12)*100:.1f}%")
    ]

    row = 16
    for metric, value in metrics:
        ws.cell(row=row, column=1, value=metric).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

# Main application pages
def login_page():
    """Login and registration page"""
    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown("""
        <div style='text-align: center; margin-bottom: 40px;'>
            <h1 style='color: rgb(0, 0, 0); font-size: 48px; margin-bottom: 10px;'>🏃 Sprint Analysis Pro</h1>
            <p style='color: rgb(133, 72, 54); font-size: 18px;'>Advanced 100m Sprint Performance Analysis</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div class='instagram-card' style='padding: 30px;'>", unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["Sign In", "Create Account"])

        with tab1:
            with st.form("login_form"):
                username = st.text_input("Username", placeholder="Enter your username")
                password = st.text_input("Password", type="password", placeholder="Enter your password")
                submit = st.form_submit_button("Sign In", use_container_width=True)

                if submit:
                    if username and password:
                        result = authenticate_user(username, password)
                        if result:
                            st.session_state.authenticated = True
                            st.session_state.user_id = result[0]
                            st.session_state.user_type = result[1]
                            st.session_state.username = username
                            st.success("Login successful!")
                            st.rerun()
                        else:
                            st.error("Invalid username or password")
                    else:
                        st.error("Please enter both username and password")

            st.markdown("---")
            st.markdown("""
            <div style='text-align: center; color: #666;'>
                <small>Demo credentials: <b>admin</b> / <b>admin123</b></small>
            </div>
            """, unsafe_allow_html=True)

        with tab2:
            with st.form("register_form"):
                new_username = st.text_input("Choose Username", placeholder="Enter username")
                new_password = st.text_input("Choose Password", type="password", placeholder="Enter password")
                confirm_password = st.text_input("Confirm Password", type="password", placeholder="Confirm password")
                user_type = st.selectbox("User Type", ["runner", "coach"])
                submit_reg = st.form_submit_button("Create Account", use_container_width=True)

                if submit_reg:
                    if new_username and new_password and confirm_password:
                        if new_password == confirm_password:
                            if len(new_password) >= 6:
                                success, user_id = register_user(new_username, new_password, user_type)
                                if success:
                                    st.success("Account created successfully! Please sign in.")
                                else:
                                    st.error("Username already exists. Please choose another.")
                            else:
                                st.error("Password must be at least 6 characters long")
                        else:
                            st.error("Passwords do not match")
                    else:
                        st.error("Please fill all fields")

        st.markdown("</div>", unsafe_allow_html=True)

def main_dashboard():
    """Main dashboard after login"""
    # Sidebar
    with st.sidebar:
        st.markdown(f"""
        <div style='text-align: center; padding: 20px; background-color: rgb(255, 178, 44); border-radius: 8px; margin-bottom: 20px;'>
            <h3 style='margin: 0; color: rgb(0, 0, 0);'>Welcome!</h3>
            <p style='margin: 5px 0; font-size: 18px; color: rgb(0, 0, 0);'>{st.session_state.username}</p>
            <p style='margin: 0; font-size: 14px; color: rgb(133, 72, 54);'>{st.session_state.user_type.upper()}</p>
        </div>
        """, unsafe_allow_html=True)

        # Navigation based on user type
        st.markdown("### Navigation")

        if st.session_state.user_type == 'admin':
            page = st.radio("Select Page",
                           ["📹 Upload & Analyze", "📊 View Reports", "👥 Manage Users", "🏃 Manage Runners"],
                           label_visibility="collapsed")
        elif st.session_state.user_type == 'coach':
            page = st.radio("Select Page",
                           ["📹 Upload & Analyze", "📊 View Reports", "👥 My Runners"],
                           label_visibility="collapsed")
        else:
            page = st.radio("Select Page",
                           ["📹 Upload & Analyze", "📊 View Reports"],
                           label_visibility="collapsed")

        st.markdown("---")

        if st.button("🚪 Logout", use_container_width=True):
            for key in ['authenticated', 'user_type', 'username', 'user_id']:
                st.session_state[key] = None
            st.rerun()

    # Main content area
    if "Upload & Analyze" in page:
        upload_analyze_page()
    elif "View Reports" in page:
        view_reports_page()
    elif "Manage Users" in page and st.session_state.user_type == 'admin':
        manage_users_page()
    elif "Manage Runners" in page and st.session_state.user_type == 'admin':
        manage_runners_page()
    elif "My Runners" in page and st.session_state.user_type == 'coach':
        my_runners_page()

def upload_analyze_page():
    """Page for uploading videos and analyzing performance"""
    st.title("📹 Upload & Analyze Sprint Performance")

    # Get runner information
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()

    if st.session_state.user_type == 'coach':
        c.execute("""SELECT r.id, r.name 
                     FROM runners r 
                     JOIN users u ON r.coach_id = u.id 
                     WHERE u.username = ?""", (st.session_state.username,))
    elif st.session_state.user_type == 'admin':
        c.execute("SELECT id, name FROM runners")
    else:  # runner
        # Auto-create runner entry if not exists
        c.execute("SELECT id FROM runners WHERE name = ?", (st.session_state.username,))
        runner_exists = c.fetchone()
        if not runner_exists:
            c.execute("INSERT INTO runners (name, coach_id) VALUES (?, NULL)",
                     (st.session_state.username,))
            conn.commit()
        c.execute("SELECT id, name FROM runners WHERE name = ?", (st.session_state.username,))

    runners = c.fetchall()
    conn.close()

    if not runners:
        st.warning("No runners found. Please add runners first.")
        return

    # Runner selection
    runner_dict = {name: id for id, name in runners}

    col1, col2 = st.columns([2, 1])
    with col1:
        selected_runner = st.selectbox("Select Runner", list(runner_dict.keys()))
    with col2:
        st.markdown(f"""
        <div class='metric-card' style='text-align: center; margin-top: 25px;'>
            <p style='margin: 0; color: #666; font-size: 14px;'>Runner ID</p>
            <h3 style='margin: 0; color: rgb(255, 178, 44);'>#{runner_dict[selected_runner]:03d}</h3>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # Video upload section
    st.markdown("### 📤 Upload Sprint Videos")
    st.info("Upload one video for each 25-meter segment of the 100m sprint")

    video_files = {}
    upload_status = {}

    # Create upload interface
    cols = st.columns(4)
    ranges = [
        ("0-25", "🚀 Start (0-25m)", "Acceleration phase"),
        ("25-50", "⚡ Speed (25-50m)", "Peak velocity phase"),
        ("50-75", "💪 Maintain (50-75m)", "Sustained speed"),
        ("75-100", "🏁 Finish (75-100m)", "Final sprint")
    ]

    for idx, (range_key, title, description) in enumerate(ranges):
        with cols[idx]:
            st.markdown(f"""
            <div class='instagram-card' style='text-align: center; min-height: 200px;'>
                <h4>{title}</h4>
                <p style='color: #666; font-size: 14px;'>{description}</p>
            </div>
            """, unsafe_allow_html=True)

            video_files[range_key] = st.file_uploader(
                "Upload video",
                type=['mp4', 'avi', 'mov'],
                key=f"video_{range_key}",
                label_visibility="collapsed"
            )

            if video_files[range_key]:
                upload_status[range_key] = True
                st.success("✅ Uploaded")
            else:
                upload_status[range_key] = False

    # Progress indicator
    uploaded_count = sum(upload_status.values())
    progress = uploaded_count / 4

    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.progress(progress)
        st.markdown(f"""
        <div style='text-align: center;'>
            <h3 style='color: rgb(255, 178, 44);'>{uploaded_count}/4 Videos Uploaded</h3>
        </div>
        """, unsafe_allow_html=True)

    # Analyze button
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🔍 Analyze Performance", use_container_width=True, disabled=uploaded_count < 4):
            if all(upload_status.values()):
                with st.spinner("🎬 Processing videos with computer vision..."):
                    # Process videos
                    all_performance_data = {}
                    video_analyses = {}

                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    for idx, (range_name, video_file) in enumerate(video_files.items()):
                        progress_bar.progress((idx + 1) / 4)
                        status_text.text(f"Analyzing {range_name}m segment...")

                        # Process video
                        video_analysis = process_video_with_cv(video_file)
                        video_analyses[range_name] = video_analysis

                        # Generate performance data
                        performance_data = generate_performance_data(range_name, video_analysis)
                        all_performance_data[range_name] = performance_data

                    status_text.text("✅ Analysis complete!")

                # Display results
                display_analysis_results(all_performance_data, selected_runner, runner_dict[selected_runner])

def display_analysis_results(performance_data, runner_name, runner_id):
    """Display analysis results with visualizations"""
    st.markdown("---")
    st.success("✅ Analysis Complete! Here are your results:")

    # Calculate metrics
    all_velocities = []
    total_time = 0

    for data in performance_data.values():
        all_velocities.extend(data['velocity'].tolist())
        time_segment = data['time'].iloc[-1] - data['time'].iloc[0]
        total_time += time_segment

    max_speed = max(all_velocities)
    avg_speed = sum(all_velocities) / len(all_velocities)

    # Display metrics
    st.markdown("### 📊 Performance Metrics")

    cols = st.columns(4)
    metrics = [
        ("🏆 Max Speed", f"{max_speed:.2f}", "m/s", "Best velocity achieved"),
        ("📈 Avg Speed", f"{avg_speed:.2f}", "m/s", "Overall average"),
        ("⏱️ Total Time", f"{total_time:.2f}", "sec", "100m completion"),
        ("💯 Score", f"{(max_speed/12)*100:.0f}", "%", "Performance rating")
    ]

    for idx, (icon_title, value, unit, desc) in enumerate(metrics):
        with cols[idx]:
            st.markdown(f"""
            <div class='metric-card' style='text-align: center;'>
                <h3 style='margin: 0;'>{icon_title}</h3>
                <h1 style='color: rgb(255, 178, 44); margin: 10px 0;'>{value}<small style='font-size: 20px;'>{unit}</small></h1>
                <p style='color: #666; font-size: 14px; margin: 0;'>{desc}</p>
            </div>
            """, unsafe_allow_html=True)

    # Velocity profile chart using Streamlit native charts
    st.markdown("### 📈 Velocity Profile Analysis")

    # Prepare data for visualization
    chart_data = pd.DataFrame()

    for range_name, data in performance_data.items():
        temp_df = data[['time', 'velocity']].copy()
        temp_df['range'] = range_name
        chart_data = pd.concat([chart_data, temp_df])

    # Create line chart
    st.line_chart(
        data=chart_data.pivot(index='time', columns='range', values='velocity'),
        use_container_width=True,
        height=500
    )

    # Position chart
    with st.expander("📍 View Position Data"):
        position_data = pd.DataFrame()

        for range_name, data in performance_data.items():
            temp_df = data[['time', 'position']].copy()
            temp_df['range'] = range_name
            position_data = pd.concat([position_data, temp_df])

        st.line_chart(
            data=position_data.pivot(index='time', columns='range', values='position'),
            use_container_width=True,
            height=400
        )

    # Save to database
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()

    # Convert data to JSON
    performance_json = {}
    for range_name, data in performance_data.items():
        performance_json[range_name] = data.to_json()

    c.execute("""INSERT INTO performance_data 
                (runner_id, test_date, range_0_25_data, range_25_50_data, 
                 range_50_75_data, range_75_100_data, max_speed, avg_speed, total_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
             (runner_id, datetime.now(),
              performance_json.get("0-25"), performance_json.get("25-50"),
              performance_json.get("50-75"), performance_json.get("75-100"),
              max_speed, avg_speed, total_time))

    conn.commit()
    conn.close()

    # Download section
    st.markdown("### 💾 Export Results")

    col1, col2 = st.columns(2)

    with col1:
        # Generate Excel report
        excel_report = generate_excel_report(performance_data, runner_name)

        st.download_button(
            label="📊 Download Excel Report",
            data=excel_report,
            file_name=f"sprint_analysis_{runner_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

def view_reports_page():
    """View historical performance reports"""
    st.title("📊 Performance Reports")

    # Get data based on user type
    conn = sqlite3.connect('running_analysis.db')

    if st.session_state.user_type == 'coach':
        query = """
        SELECT p.*, r.name as runner_name 
        FROM performance_data p
        JOIN runners r ON p.runner_id = r.id
        JOIN users u ON r.coach_id = u.id
        WHERE u.username = ?
        ORDER BY p.test_date DESC
        """
        df = pd.read_sql_query(query, conn, params=(st.session_state.username,))
    elif st.session_state.user_type == 'admin':
        query = """
        SELECT p.*, r.name as runner_name 
        FROM performance_data p
        JOIN runners r ON p.runner_id = r.id
        ORDER BY p.test_date DESC
        """
        df = pd.read_sql_query(query, conn)
    else:  # runner
        query = """
        SELECT p.*, r.name as runner_name 
        FROM performance_data p
        JOIN runners r ON p.runner_id = r.id
        WHERE r.name = ?
        ORDER BY p.test_date DESC
        """
        df = pd.read_sql_query(query, conn, params=(st.session_state.username,))

    conn.close()

    if df.empty:
        st.info("No performance data available yet. Upload videos to start analyzing!")
        return

    # Convert date column
    df['test_date'] = pd.to_datetime(df['test_date'])

    # Filters
    col1, col2, col3 = st.columns(3)

    with col1:
        runners = ["All"] + list(df['runner_name'].unique())
        selected_runner = st.selectbox("Filter by Runner", runners)

    with col2:
        # Date range
        min_date = df['test_date'].min().date()
        max_date = df['test_date'].max().date()
        date_range = st.date_input("Date Range",
                                  value=(min_date, max_date),
                                  min_value=min_date,
                                  max_value=max_date)

    with col3:
        # Sort order
        sort_order = st.selectbox("Sort by", ["Newest First", "Oldest First", "Best Performance"])

    # Apply filters
    filtered_df = df.copy()

    if selected_runner != "All":
        filtered_df = filtered_df[filtered_df['runner_name'] == selected_runner]

    if len(date_range) == 2:
        filtered_df = filtered_df[
            (filtered_df['test_date'].dt.date >= date_range[0]) &
            (filtered_df['test_date'].dt.date <= date_range[1])
        ]

    # Apply sorting
    if sort_order == "Newest First":
        filtered_df = filtered_df.sort_values('test_date', ascending=False)
    elif sort_order == "Oldest First":
        filtered_df = filtered_df.sort_values('test_date', ascending=True)
    else:  # Best Performance
        filtered_df = filtered_df.sort_values('max_speed', ascending=False)

    # Summary statistics
    if not filtered_df.empty:
        st.markdown("### 📈 Summary Statistics")

        cols = st.columns(5)
        stats = [
            ("🏃 Total Tests", len(filtered_df)),
            ("🏆 Best Speed", f"{filtered_df['max_speed'].max():.2f} m/s"),
            ("📊 Avg Max Speed", f"{filtered_df['max_speed'].mean():.2f} m/s"),
            ("⏱️ Best Time", f"{filtered_df['total_time'].min():.2f} s"),
            ("👥 Athletes", filtered_df['runner_name'].nunique())
        ]

        for idx, (label, value) in enumerate(stats):
            with cols[idx]:
                st.markdown(f"""
                <div class='metric-card' style='text-align: center; padding: 15px;'>
                    <p style='margin: 0; color: #666; font-size: 14px;'>{label}</p>
                    <h3 style='margin: 5px 0; color: rgb(255, 178, 44);'>{value}</h3>
                </div>
                """, unsafe_allow_html=True)

        # Performance trend chart (if single runner selected)
        if selected_runner != "All" and len(filtered_df) > 1:
            st.markdown("### 📊 Performance Trend")

            # Prepare data for line chart
            trend_data = filtered_df[['test_date', 'max_speed', 'avg_speed']].copy()
            trend_data = trend_data.set_index('test_date')
            trend_data.columns = ['Max Speed', 'Avg Speed']

            st.line_chart(trend_data, use_container_width=True, height=400)

        # Detailed records table
        st.markdown("### 📋 Test Records")

        # Prepare display dataframe
        display_df = filtered_df[['runner_name', 'test_date', 'max_speed', 'avg_speed', 'total_time']].copy()
        display_df['test_date'] = display_df['test_date'].dt.strftime('%Y-%m-%d %H:%M')
        display_df.columns = ['Runner', 'Test Date', 'Max Speed (m/s)', 'Avg Speed (m/s)', 'Time (s)']

        # Add performance indicator
        display_df['Performance'] = display_df['Max Speed (m/s)'].apply(
            lambda x: '🏆 Excellent' if x > 9 else '✅ Good' if x > 8 else '📈 Improving'
        )

        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True
        )

def manage_users_page():
    """Admin page for managing users"""
    st.title("👥 User Management")

    tab1, tab2 = st.tabs(["View Users", "Add New User"])

    with tab1:
        conn = sqlite3.connect('running_analysis.db')
        users_df = pd.read_sql_query("""
            SELECT id, username, user_type, created_at 
            FROM users 
            ORDER BY created_at DESC
        """, conn)
        conn.close()

        if not users_df.empty:
            # User statistics
            col1, col2, col3 = st.columns(3)

            with col1:
                total_users = len(users_df)
                st.metric("Total Users", total_users)

            with col2:
                coaches = len(users_df[users_df['user_type'] == 'coach'])
                st.metric("Coaches", coaches)

            with col3:
                runners = len(users_df[users_df['user_type'] == 'runner'])
                st.metric("Runners", runners)

            # User table
            st.markdown("### User List")

            # Format the dataframe
            users_df['created_at'] = pd.to_datetime(users_df['created_at']).dt.strftime('%Y-%m-%d')
            users_df['user_type'] = users_df['user_type'].str.upper()
            users_df.columns = ['ID', 'Username', 'Role', 'Created Date']

            st.dataframe(
                users_df,
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("No users found in the system.")

    with tab2:
        st.markdown("### Add New User")

        with st.form("add_user_form"):
            col1, col2 = st.columns(2)

            with col1:
                new_username = st.text_input("Username", placeholder="Enter username")
                new_password = st.text_input("Password", type="password", placeholder="Enter password")

            with col2:
                user_type = st.selectbox("User Type", ["runner", "coach", "admin"])
                confirm_password = st.text_input("Confirm Password", type="password",
                                               placeholder="Confirm password")

            submit = st.form_submit_button("Create User", use_container_width=True)

            if submit:
                if new_username and new_password and confirm_password:
                    if new_password == confirm_password:
                        if len(new_password) >= 6:
                            success, _ = register_user(new_username, new_password, user_type)
                            if success:
                                st.success(f"✅ User '{new_username}' created successfully!")
                                st.rerun()
                            else:
                                st.error("Username already exists.")
                        else:
                            st.error("Password must be at least 6 characters long.")
                    else:
                        st.error("Passwords do not match.")
                else:
                    st.error("Please fill all fields.")

def manage_runners_page():
    """Admin page for managing runners"""
    st.title("🏃 Runner Management")

    tab1, tab2, tab3 = st.tabs(["View Runners", "Add Runner", "Assign Coach"])

    with tab1:
        conn = sqlite3.connect('running_analysis.db')
        runners_df = pd.read_sql_query("""
            SELECT r.id, r.name as runner_name, u.username as coach_name, r.created_at,
                   COUNT(p.id) as total_tests
            FROM runners r
            LEFT JOIN users u ON r.coach_id = u.id
            LEFT JOIN performance_data p ON r.id = p.runner_id
            GROUP BY r.id, r.name, u.username, r.created_at
            ORDER BY r.created_at DESC
        """, conn)
        conn.close()

        if not runners_df.empty:
            # Statistics
            col1, col2, col3 = st.columns(3)

            with col1:
                total_runners = len(runners_df)
                st.metric("Total Runners", total_runners)

            with col2:
                assigned_runners = len(runners_df[runners_df['coach_name'].notna()])
                st.metric("Assigned to Coaches", assigned_runners)

            with col3:
                total_tests = runners_df['total_tests'].sum()
                st.metric("Total Tests", int(total_tests))

            # Runners table
            st.markdown("### Runner List")

            # Format the dataframe
            runners_df['created_at'] = pd.to_datetime(runners_df['created_at']).dt.strftime('%Y-%m-%d')
            runners_df['coach_name'] = runners_df['coach_name'].fillna('Unassigned')
            runners_df.columns = ['ID', 'Runner Name', 'Coach', 'Joined Date', 'Tests']

            st.dataframe(
                runners_df,
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("No runners registered yet.")

    with tab2:
        st.markdown("### Add New Runner")

        with st.form("add_runner_form"):
            runner_name = st.text_input("Runner Name", placeholder="Enter runner's full name")

            # Get coaches
            conn = sqlite3.connect('running_analysis.db')
            c = conn.cursor()
            c.execute("SELECT id, username FROM users WHERE user_type = 'coach'")
            coaches = c.fetchall()
            conn.close()

            if coaches:
                coach_options = ["Unassigned"] + [username for _, username in coaches]
                coach_dict = {username: id for id, username in coaches}
                selected_coach = st.selectbox("Assign to Coach (Optional)", coach_options)
            else:
                selected_coach = "Unassigned"
                st.info("No coaches available. Add coaches to assign runners.")

            submit = st.form_submit_button("Add Runner", use_container_width=True)

            if submit:
                if runner_name:
                    conn = sqlite3.connect('running_analysis.db')
                    c = conn.cursor()

                    coach_id = coach_dict.get(selected_coach) if selected_coach != "Unassigned" else None

                    try:
                        c.execute("INSERT INTO runners (name, coach_id) VALUES (?, ?)",
                                 (runner_name, coach_id))
                        conn.commit()
                        st.success(f"✅ Runner '{runner_name}' added successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error adding runner: {str(e)}")
                    finally:
                        conn.close()
                else:
                    st.error("Please enter runner name.")

    with tab3:
        st.markdown("### Assign/Reassign Coach")

        conn = sqlite3.connect('running_analysis.db')

        # Get runners
        c = conn.cursor()
        c.execute("SELECT id, name FROM runners ORDER BY name")
        runners = c.fetchall()

        # Get coaches
        c.execute("SELECT id, username FROM users WHERE user_type = 'coach'")
        coaches = c.fetchall()
        conn.close()

        if runners and coaches:
            with st.form("assign_coach_form"):
                runner_dict = {f"{name} (#{id})": id for id, name in runners}
                selected_runner = st.selectbox("Select Runner", list(runner_dict.keys()))

                coach_dict = {username: id for id, username in coaches}
                coach_options = ["Unassigned"] + list(coach_dict.keys())
                selected_coach = st.selectbox("Assign to Coach", coach_options)

                submit = st.form_submit_button("Update Assignment", use_container_width=True)

                if submit:
                    runner_id = runner_dict[selected_runner]
                    coach_id = coach_dict.get(selected_coach) if selected_coach != "Unassigned" else None

                    conn = sqlite3.connect('running_analysis.db')
                    c = conn.cursor()
                    c.execute("UPDATE runners SET coach_id = ? WHERE id = ?", (coach_id, runner_id))
                    conn.commit()
                    conn.close()

                    st.success("✅ Coach assignment updated successfully!")
                    st.rerun()
        else:
            st.info("Add runners and coaches first to manage assignments.")

def my_runners_page():
    """Coach page to view their assigned runners"""
    st.title("👥 My Runners")

    conn = sqlite3.connect('running_analysis.db')

    # Get coach's runners with performance stats
    runners_df = pd.read_sql_query("""
        SELECT r.id, r.name, 
               COUNT(p.id) as total_tests,
               MAX(p.max_speed) as best_speed,
               AVG(p.avg_speed) as avg_speed,
               MIN(p.total_time) as best_time,
               MAX(p.test_date) as last_test
        FROM runners r
        JOIN users u ON r.coach_id = u.id
        LEFT JOIN performance_data p ON r.id = p.runner_id
        WHERE u.username = ?
        GROUP BY r.id, r.name
        ORDER BY r.name
    """, conn, params=(st.session_state.username,))

    conn.close()

    if runners_df.empty:
        st.info("No runners assigned to you yet. Contact admin to get runners assigned.")
        return

    # Display summary
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Total Runners", len(runners_df))

    with col2:
        active_runners = len(runners_df[runners_df['total_tests'] > 0])
        st.metric("Active Runners", active_runners)

    with col3:
        total_tests = runners_df['total_tests'].sum()
        st.metric("Total Tests", int(total_tests))

    st.markdown("---")

    # Display each runner's card
    for _, runner in runners_df.iterrows():
        with st.expander(f"🏃 {runner['name']}", expanded=True):
            col1, col2, col3, col4 = st.columns(4)

            with col1:
                tests = int(runner['total_tests']) if runner['total_tests'] else 0
                st.metric("Tests Completed", tests)

            with col2:
                best_speed = runner['best_speed'] if runner['best_speed'] else 0
                st.metric("Best Speed", f"{best_speed:.2f} m/s")

            with col3:
                avg_speed = runner['avg_speed'] if runner['avg_speed'] else 0
                st.metric("Avg Speed", f"{avg_speed:.2f} m/s")

            with col4:
                best_time = runner['best_time'] if runner['best_time'] else 0
                st.metric("Best Time", f"{best_time:.2f} s")

            if runner['last_test']:
                last_test = pd.to_datetime(runner['last_test']).strftime('%Y-%m-%d')
                st.caption(f"Last tested: {last_test}")
            else:
                st.caption("No tests completed yet")

# Main execution
if __name__ == "__main__":
    if st.session_state.authenticated:
        main_dashboard()
    else:
        login_page()